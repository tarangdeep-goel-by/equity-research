"""Screener.in Excel export client for historical quarterly results backfill."""

from __future__ import annotations

import io
import re
from datetime import datetime
from pathlib import Path

import httpx
import openpyxl

from bs4 import BeautifulSoup

from flowtracker.fund_models import AnnualEPS, AnnualFinancials, QuarterlyResult, ScreenerRatios

_SCREENER_BASE = "https://www.screener.in"
_CRED_PATH = Path.home() / ".config" / "flowtracker" / "screener.env"


class ScreenerError(Exception):
    """Raised on Screener.in fetch failures."""


def _load_credentials() -> tuple[str, str]:
    """Load email and password from ~/.config/flowtracker/screener.env"""
    if not _CRED_PATH.exists():
        raise ScreenerError(
            f"Screener.in credentials not found at {_CRED_PATH}\n"
            "Create the file with:\n"
            "  SCREENER_EMAIL=your@email.com\n"
            "  SCREENER_PASSWORD=yourpassword"
        )
    creds = {}
    for line in _CRED_PATH.read_text().strip().splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        key, _, val = line.partition("=")
        creds[key.strip()] = val.strip()
    email = creds.get("SCREENER_EMAIL", "")
    password = creds.get("SCREENER_PASSWORD", "")
    if not email or not password:
        raise ScreenerError("SCREENER_EMAIL and SCREENER_PASSWORD must be set in screener.env")
    return email, password


def _parse_screener_date(date_str: str) -> str:
    """Convert Screener.in date header (e.g., 'Dec 2025', 'Mar 2016') to ISO quarter-end date.

    Screener.in uses month abbreviations. Map to quarter-end dates:
    Mar YYYY -> YYYY-03-31
    Jun YYYY -> YYYY-06-30
    Sep YYYY -> YYYY-09-30
    Dec YYYY -> YYYY-12-31
    """
    date_str = date_str.strip()
    try:
        dt = datetime.strptime(date_str, "%b %Y")
    except ValueError:
        # Try other formats
        try:
            dt = datetime.strptime(date_str, "%B %Y")
        except ValueError:
            raise ScreenerError(f"Cannot parse date: {date_str}")

    # Map to quarter-end
    month = dt.month
    year = dt.year
    quarter_ends = {3: "03-31", 6: "06-30", 9: "09-30", 12: "12-31"}
    if month in quarter_ends:
        return f"{year}-{quarter_ends[month]}"
    # Non-quarter month — use last day of that month
    import calendar

    last_day = calendar.monthrange(year, month)[1]
    return f"{year}-{month:02d}-{last_day:02d}"


class ScreenerClient:
    """Screener.in client for downloading Excel exports with 10yr quarterly data."""

    def __init__(self) -> None:
        email, password = _load_credentials()
        self._client = httpx.Client(
            follow_redirects=True,
            timeout=30,
            headers={"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"},
        )
        self._login(email, password)

    def _login(self, email: str, password: str) -> None:
        """Login to Screener.in and establish session cookies."""
        # Step 1: GET login page to get CSRF token
        resp = self._client.get(f"{_SCREENER_BASE}/login/")
        resp.raise_for_status()
        csrf = self._client.cookies.get("csrftoken")
        if not csrf:
            raise ScreenerError("Failed to get CSRF token from login page")

        # Step 2: POST login
        resp = self._client.post(
            f"{_SCREENER_BASE}/login/",
            data={
                "csrfmiddlewaretoken": csrf,
                "username": email,
                "password": password,
            },
            headers={"Referer": f"{_SCREENER_BASE}/login/"},
        )
        resp.raise_for_status()

        # Check if login succeeded (Django redirects to / on success, back to /login/ on failure)
        if "/login/" in str(resp.url):
            raise ScreenerError("Login failed — check credentials in screener.env")

    def _get_warehouse_id(self, symbol: str) -> str:
        """Fetch company page and extract the Excel export warehouse ID."""
        html = self.fetch_company_page(symbol)
        match = re.search(r'formaction="/user/company/export/(\d+)/"', html)
        if not match:
            match = re.search(r"/user/company/export/(\d+)/", html)
        if not match:
            raise ScreenerError(f"Could not find export warehouse ID for {symbol}")
        return match.group(1)

    def fetch_company_page(self, symbol: str) -> str:
        """Fetch Screener.in company page HTML (consolidated preferred)."""
        for suffix in ["/consolidated/", "/"]:
            url = f"{_SCREENER_BASE}/company/{symbol}{suffix}"
            resp = self._client.get(url)
            if resp.status_code == 200:
                self._last_html = resp.text
                return resp.text
        raise ScreenerError(f"Company page not found for {symbol}")

    def download_excel(self, symbol: str) -> bytes:
        """Download the Excel export for a symbol."""
        warehouse_id = self._get_warehouse_id(symbol)
        csrf = self._client.cookies.get("csrftoken")

        resp = self._client.post(
            f"{_SCREENER_BASE}/user/company/export/{warehouse_id}/",
            data={"csrfmiddlewaretoken": csrf},
            headers={"Referer": f"{_SCREENER_BASE}/company/{symbol}/consolidated/"},
        )
        resp.raise_for_status()

        if resp.headers.get("content-type", "").startswith("text/html"):
            raise ScreenerError(f"Export failed for {symbol} — got HTML instead of Excel")

        return resp.content

    def parse_quarterly_results(
        self, symbol: str, excel_bytes: bytes
    ) -> list[QuarterlyResult]:
        """Parse quarterly data from the 'Data Sheet' in a Screener.in Excel export.

        The 'Quarters' sheet uses formulas referencing 'Data Sheet', so we read
        the raw data directly from 'Data Sheet' where the Quarters section starts
        with a row containing ('Quarters', None, ...) followed by 'Report Date'
        and metric rows (Sales, Expenses, Operating Profit, etc.).
        """
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)

        if "Data Sheet" not in wb.sheetnames:
            wb.close()
            raise ScreenerError(f"No 'Data Sheet' found in export for {symbol}")

        ws = wb["Data Sheet"]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Find the Quarters section: row where first cell == "Quarters"
        q_start = None
        for i, row in enumerate(all_rows):
            if row and isinstance(row[0], str) and row[0].strip().lower() == "quarters":
                q_start = i
                break
        if q_start is None:
            raise ScreenerError(f"No 'Quarters' section found in Data Sheet for {symbol}")

        # Next row should be Report Date with datetime objects
        date_row = all_rows[q_start + 1] if q_start + 1 < len(all_rows) else None
        if not date_row or str(date_row[0]).strip().lower() != "report date":
            raise ScreenerError(f"Expected 'Report Date' row after Quarters header for {symbol}")

        # Extract quarter-end dates from date columns
        date_cols: list[tuple[int, str]] = []  # (col_index, ISO date string)
        for col_idx in range(1, len(date_row)):
            val = date_row[col_idx]
            if val is None:
                continue
            if isinstance(val, datetime):
                date_cols.append((col_idx, val.strftime("%Y-%m-%d")))
            elif isinstance(val, str):
                try:
                    date_cols.append((col_idx, _parse_screener_date(val)))
                except ScreenerError:
                    continue

        if not date_cols:
            return []

        # Collect metric rows until we hit an empty row or a new section header
        metrics: dict[str, list[float | None]] = {}
        for row in all_rows[q_start + 2 :]:
            if not row or row[0] is None:
                break  # End of Quarters section
            label = str(row[0]).strip().lower()
            if not label:
                break
            values: list[float | None] = []
            for col_idx, _ in date_cols:
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        val = None
                values.append(val)
            metrics[label] = values

        # Helper to look up a metric by possible label names
        def _get(keys: list[str]) -> list[float | None]:
            for key in keys:
                if key in metrics:
                    return metrics[key]
            return [None] * len(date_cols)

        revenue_vals = _get(["sales", "revenue", "total revenue"])
        expenses_vals = _get(["expenses"])
        op_income_vals = _get(["operating profit"])
        opm_vals = _get(["opm %", "opm%"])
        other_income_vals = _get(["other income"])
        interest_vals = _get(["interest"])
        depreciation_vals = _get(["depreciation"])
        pbt_vals = _get(["profit before tax"])
        tax_pct_vals = _get(["tax %", "tax%"])
        net_income_vals = _get(["net profit", "profit after tax", "pat"])
        eps_vals = _get(["eps in rs", "eps"])

        # Build QuarterlyResult for each quarter
        results = []
        for i, (_, quarter_end) in enumerate(date_cols):
            revenue = revenue_vals[i] if i < len(revenue_vals) else None
            operating_income = op_income_vals[i] if i < len(op_income_vals) else None
            net_income = net_income_vals[i] if i < len(net_income_vals) else None
            depreciation = depreciation_vals[i] if i < len(depreciation_vals) else None

            # EBITDA = operating_income + depreciation
            ebitda = None
            if operating_income is not None and depreciation is not None:
                ebitda = operating_income + depreciation

            # Use Screener's OPM% directly (as fraction)
            opm_raw = opm_vals[i] if i < len(opm_vals) else None
            operating_margin = opm_raw / 100 if opm_raw is not None else None
            if operating_margin is None and operating_income is not None and revenue and revenue != 0:
                operating_margin = operating_income / revenue

            net_margin = None
            if net_income is not None and revenue and revenue != 0:
                net_margin = net_income / revenue

            results.append(
                QuarterlyResult(
                    symbol=symbol,
                    quarter_end=quarter_end,
                    revenue=revenue,
                    expenses=expenses_vals[i] if i < len(expenses_vals) else None,
                    operating_income=operating_income,
                    net_income=net_income,
                    ebitda=ebitda,
                    eps=eps_vals[i] if i < len(eps_vals) else None,
                    operating_margin=operating_margin,
                    net_margin=net_margin,
                    other_income=other_income_vals[i] if i < len(other_income_vals) else None,
                    depreciation=depreciation,
                    interest=interest_vals[i] if i < len(interest_vals) else None,
                    profit_before_tax=pbt_vals[i] if i < len(pbt_vals) else None,
                    tax_pct=tax_pct_vals[i] if i < len(tax_pct_vals) else None,
                )
            )

        # Sort by quarter_end ascending (oldest first)
        results.sort(key=lambda r: r.quarter_end)
        return results

    def parse_annual_eps(self, symbol: str, excel_bytes: bytes) -> list[AnnualEPS]:
        """Parse annual EPS from the P&L section of the Data Sheet.

        Returns ~10 years of annual EPS for historical P/E computation.
        """
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)

        if "Data Sheet" not in wb.sheetnames:
            wb.close()
            raise ScreenerError(f"No 'Data Sheet' found in export for {symbol}")

        ws = wb["Data Sheet"]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Find PROFIT & LOSS section
        pl_start = None
        for i, row in enumerate(all_rows):
            if row and isinstance(row[0], str) and row[0].strip().upper() == "PROFIT & LOSS":
                pl_start = i
                break
        if pl_start is None:
            return []

        # Next row should be Report Date
        date_row = all_rows[pl_start + 1] if pl_start + 1 < len(all_rows) else None
        if not date_row:
            return []

        # Extract fiscal year end dates
        date_cols: list[tuple[int, str]] = []
        for col_idx in range(1, len(date_row)):
            val = date_row[col_idx]
            if val is None:
                continue
            if isinstance(val, datetime):
                date_cols.append((col_idx, val.strftime("%Y-%m-%d")))
            elif isinstance(val, str):
                val_str = val.strip()
                if not val_str or val_str.lower() in ("ttm", "report date"):
                    continue
                try:
                    date_cols.append((col_idx, _parse_screener_date(val_str)))
                except ScreenerError:
                    continue

        if not date_cols:
            return []

        # Find the Quarters section start to know where P&L ends
        q_start = None
        for i, row in enumerate(all_rows):
            if row and isinstance(row[0], str) and row[0].strip().lower() == "quarters":
                q_start = i
                break
        pl_end = q_start if q_start else len(all_rows)

        # Collect metric rows from P&L section
        metrics: dict[str, list[float | None]] = {}
        for row in all_rows[pl_start + 2 : pl_end]:
            if not row or row[0] is None:
                continue
            label = str(row[0]).strip().lower()
            if not label:
                continue
            values: list[float | None] = []
            for col_idx, _ in date_cols:
                val = row[col_idx] if col_idx < len(row) else None
                if val is not None:
                    try:
                        val = float(val)
                    except (ValueError, TypeError):
                        val = None
                values.append(val)
            metrics[label] = values

        def _get(keys: list[str]) -> list[float | None]:
            for key in keys:
                if key in metrics:
                    return metrics[key]
            return [None] * len(date_cols)

        revenue_vals = _get(["sales", "revenue", "total revenue"])
        net_income_vals = _get(["net profit", "profit after tax", "pat"])

        # EPS not directly in P&L section; find shares row after P&L section
        # Row labeled "Adjusted Equity Shares in Cr" is typically near end of Data Sheet
        shares_vals: list[float | None] = [None] * len(date_cols)
        for row in all_rows:
            if row and isinstance(row[0], str) and "adjusted equity shares" in row[0].strip().lower():
                for j, (col_idx, _) in enumerate(date_cols):
                    val = row[col_idx] if col_idx < len(row) else None
                    if val is not None:
                        try:
                            shares_vals[j] = float(val)
                        except (ValueError, TypeError):
                            pass
                break

        results = []
        for i, (_, fy_end) in enumerate(date_cols):
            net_income = net_income_vals[i] if i < len(net_income_vals) else None
            shares = shares_vals[i] if i < len(shares_vals) else None

            # Compute EPS = Net Profit (Cr) / Adjusted Shares (Cr)
            eps = None
            if net_income is not None and shares and shares > 0:
                eps = net_income / shares

            if eps is None:
                continue  # Skip years without computable EPS
            results.append(AnnualEPS(
                symbol=symbol,
                fiscal_year_end=fy_end,
                eps=eps,
                revenue=revenue_vals[i] if i < len(revenue_vals) else None,
                net_income=net_income,
            ))

        results.sort(key=lambda r: r.fiscal_year_end)
        return results

    def parse_annual_financials(self, symbol: str, excel_bytes: bytes) -> list[AnnualFinancials]:
        """Parse full annual financials from all sections of the Data Sheet."""
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=False)
        if "Data Sheet" not in wb.sheetnames:
            wb.close()
            return []
        ws = wb["Data Sheet"]
        all_rows = list(ws.iter_rows(values_only=True))
        wb.close()

        # Find section start rows
        sections = {}
        for i, row in enumerate(all_rows):
            if row and isinstance(row[0], str):
                label = row[0].strip().upper()
                if label in ("PROFIT & LOSS", "QUARTERS", "BALANCE SHEET", "CASH FLOW:", "PRICE:", "DERIVED:"):
                    sections[label] = i

        if "PROFIT & LOSS" not in sections:
            return []

        # Get date columns from P&L Report Date row
        pl_start = sections["PROFIT & LOSS"]
        date_row = all_rows[pl_start + 1] if pl_start + 1 < len(all_rows) else None
        if not date_row:
            return []

        date_cols: list[tuple[int, str]] = []
        for col_idx in range(1, len(date_row)):
            val = date_row[col_idx]
            if val is None:
                continue
            if isinstance(val, datetime):
                date_cols.append((col_idx, val.strftime("%Y-%m-%d")))
            elif isinstance(val, str):
                val_str = val.strip()
                if not val_str or val_str.lower() in ("ttm", "report date"):
                    continue
                try:
                    date_cols.append((col_idx, _parse_screener_date(val_str)))
                except ScreenerError:
                    continue

        if not date_cols:
            return []

        # Helper: extract a row's values aligned to date_cols
        def _extract_row_values(section_start: int, section_end: int, label_match: str) -> list[float | None]:
            """Find a row by label within a section and extract its values."""
            for row in all_rows[section_start:section_end]:
                if row and isinstance(row[0], str) and row[0].strip().lower() == label_match.lower():
                    values = []
                    for col_idx, _ in date_cols:
                        val = row[col_idx] if col_idx < len(row) else None
                        if val is not None:
                            try:
                                val = float(val)
                            except (ValueError, TypeError):
                                val = None
                        values.append(val)
                    return values
            return [None] * len(date_cols)

        # Section boundaries
        pl_end = sections.get("QUARTERS", sections.get("BALANCE SHEET", len(all_rows)))
        bs_start = sections.get("BALANCE SHEET", len(all_rows))
        bs_end = sections.get("CASH FLOW:", len(all_rows))
        cf_start = sections.get("CASH FLOW:", len(all_rows))
        cf_end = sections.get("PRICE:", len(all_rows))
        price_start = sections.get("PRICE:", len(all_rows))
        derived_start = sections.get("DERIVED:", len(all_rows))

        # P&L fields
        revenue = _extract_row_values(pl_start, pl_end, "Sales")
        employee_cost = _extract_row_values(pl_start, pl_end, "Employee Cost")
        raw_material_cost = _extract_row_values(pl_start, pl_end, "Raw Material Cost")
        power_and_fuel = _extract_row_values(pl_start, pl_end, "Power & Fuel Cost")
        other_mfr_exp = _extract_row_values(pl_start, pl_end, "Other Manufacturing Expenses")
        selling_and_admin = _extract_row_values(pl_start, pl_end, "Selling and Admin Expenses")
        other_expenses_detail = _extract_row_values(pl_start, pl_end, "Other Expenses")
        total_expenses = _extract_row_values(pl_start, pl_end, "Expenses")
        operating_profit = _extract_row_values(pl_start, pl_end, "Operating Profit")
        other_income = _extract_row_values(pl_start, pl_end, "Other Income")
        depreciation = _extract_row_values(pl_start, pl_end, "Depreciation")
        interest = _extract_row_values(pl_start, pl_end, "Interest")
        pbt = _extract_row_values(pl_start, pl_end, "Profit before tax")
        tax = _extract_row_values(pl_start, pl_end, "Tax")
        net_income = _extract_row_values(pl_start, pl_end, "Net profit")
        dividend = _extract_row_values(pl_start, pl_end, "Dividend Amount")

        # Balance Sheet fields
        equity_capital = _extract_row_values(bs_start, bs_end, "Equity Share Capital")
        reserves = _extract_row_values(bs_start, bs_end, "Reserves")
        borrowings = _extract_row_values(bs_start, bs_end, "Borrowings")
        other_liabilities = _extract_row_values(bs_start, bs_end, "Other Liabilities")
        total_assets_vals = []
        # "Total" appears twice in BS (liabilities and assets side). Take first one.
        found_total = False
        for row in all_rows[bs_start:bs_end]:
            if row and isinstance(row[0], str) and row[0].strip() == "Total" and not found_total:
                found_total = True
                vals = []
                for col_idx, _ in date_cols:
                    val = row[col_idx] if col_idx < len(row) else None
                    if val is not None:
                        try:
                            val = float(val)
                        except (ValueError, TypeError):
                            val = None
                    vals.append(val)
                total_assets_vals = vals
                break
        if not total_assets_vals:
            total_assets_vals = [None] * len(date_cols)

        net_block = _extract_row_values(bs_start, bs_end, "Net Block")
        cwip = _extract_row_values(bs_start, bs_end, "Capital Work in Progress")
        investments = _extract_row_values(bs_start, bs_end, "Investments")
        other_assets = _extract_row_values(bs_start, bs_end, "Other Assets")
        receivables = _extract_row_values(bs_start, bs_end, "Receivables")
        inventory = _extract_row_values(bs_start, bs_end, "Inventory")
        cash_bank = _extract_row_values(bs_start, bs_end, "Cash & Bank")
        num_shares_vals = _extract_row_values(bs_start, bs_end, "No. of Equity Shares")

        # Cash Flow fields
        cfo = _extract_row_values(cf_start, cf_end, "Cash from Operating Activity")
        cfi = _extract_row_values(cf_start, cf_end, "Cash from Investing Activity")
        cff = _extract_row_values(cf_start, cf_end, "Cash from Financing Activity")
        net_cf = _extract_row_values(cf_start, cf_end, "Net Cash Flow")

        # Price row (not inside a labeled section — it's just a row with label "PRICE:")
        price_vals = [None] * len(date_cols)
        if price_start < len(all_rows):
            row = all_rows[price_start]
            if row:
                for j, (col_idx, _) in enumerate(date_cols):
                    val = row[col_idx] if col_idx < len(row) else None
                    if val is not None:
                        try:
                            price_vals[j] = float(val)
                        except (ValueError, TypeError):
                            pass

        # EPS from adjusted shares
        adj_shares = [None] * len(date_cols)
        if derived_start < len(all_rows):
            for row in all_rows[derived_start:]:
                if row and isinstance(row[0], str) and "adjusted equity shares" in row[0].strip().lower():
                    for j, (col_idx, _) in enumerate(date_cols):
                        val = row[col_idx] if col_idx < len(row) else None
                        if val is not None:
                            try:
                                adj_shares[j] = float(val)
                            except (ValueError, TypeError):
                                pass
                    break

        # Build AnnualFinancials objects
        results = []
        for i, (_, fy_end) in enumerate(date_cols):
            ni = net_income[i] if i < len(net_income) else None
            shares = adj_shares[i] if i < len(adj_shares) else None
            eps_val = (ni / shares) if ni is not None and shares and shares > 0 else None

            results.append(AnnualFinancials(
                symbol=symbol,
                fiscal_year_end=fy_end,
                revenue=revenue[i] if i < len(revenue) else None,
                employee_cost=employee_cost[i] if i < len(employee_cost) else None,
                raw_material_cost=raw_material_cost[i] if i < len(raw_material_cost) else None,
                power_and_fuel=power_and_fuel[i] if i < len(power_and_fuel) else None,
                other_mfr_exp=other_mfr_exp[i] if i < len(other_mfr_exp) else None,
                selling_and_admin=selling_and_admin[i] if i < len(selling_and_admin) else None,
                other_expenses_detail=other_expenses_detail[i] if i < len(other_expenses_detail) else None,
                total_expenses=total_expenses[i] if i < len(total_expenses) else None,
                operating_profit=operating_profit[i] if i < len(operating_profit) else None,
                other_income=other_income[i] if i < len(other_income) else None,
                depreciation=depreciation[i] if i < len(depreciation) else None,
                interest=interest[i] if i < len(interest) else None,
                profit_before_tax=pbt[i] if i < len(pbt) else None,
                tax=tax[i] if i < len(tax) else None,
                net_income=ni,
                eps=eps_val,
                dividend_amount=dividend[i] if i < len(dividend) else None,
                equity_capital=equity_capital[i] if i < len(equity_capital) else None,
                reserves=reserves[i] if i < len(reserves) else None,
                borrowings=borrowings[i] if i < len(borrowings) else None,
                other_liabilities=other_liabilities[i] if i < len(other_liabilities) else None,
                total_assets=total_assets_vals[i] if i < len(total_assets_vals) else None,
                net_block=net_block[i] if i < len(net_block) else None,
                cwip=cwip[i] if i < len(cwip) else None,
                investments=investments[i] if i < len(investments) else None,
                other_assets=other_assets[i] if i < len(other_assets) else None,
                receivables=receivables[i] if i < len(receivables) else None,
                inventory=inventory[i] if i < len(inventory) else None,
                cash_and_bank=cash_bank[i] if i < len(cash_bank) else None,
                num_shares=num_shares_vals[i] if i < len(num_shares_vals) else None,
                cfo=cfo[i] if i < len(cfo) else None,
                cfi=cfi[i] if i < len(cfi) else None,
                cff=cff[i] if i < len(cff) else None,
                net_cash_flow=net_cf[i] if i < len(net_cf) else None,
                price=price_vals[i] if i < len(price_vals) else None,
            ))

        results.sort(key=lambda r: r.fiscal_year_end)
        return results

    def fetch_all(self, symbol: str) -> list[QuarterlyResult]:
        """Download Excel and parse quarterly results in one call."""
        excel_bytes = self.download_excel(symbol)
        return self.parse_quarterly_results(symbol, excel_bytes)

    def fetch_annual_eps(self, symbol: str) -> list[AnnualEPS]:
        """Download Excel and parse annual EPS in one call."""
        excel_bytes = self.download_excel(symbol)
        return self.parse_annual_eps(symbol, excel_bytes)

    def fetch_all_with_annual(self, symbol: str) -> tuple[list[QuarterlyResult], list[AnnualEPS], list[AnnualFinancials]]:
        """Download once, parse quarterly results, annual EPS, and full annual financials."""
        excel_bytes = self.download_excel(symbol)
        quarters = self.parse_quarterly_results(symbol, excel_bytes)
        annual_eps = self.parse_annual_eps(symbol, excel_bytes)
        annual_fin = self.parse_annual_financials(symbol, excel_bytes)
        return quarters, annual_eps, annual_fin

    # -- HTML Parsing --

    @staticmethod
    def _parse_table_section(soup: BeautifulSoup, section_id: str) -> dict[str, list[tuple[str, float | None]]]:
        """Parse a Screener data table section into {row_label: [(date, value), ...]}."""
        section = soup.find("section", id=section_id)
        if not section:
            return {}

        table = section.find("table", class_="data-table")
        if not table:
            return {}

        # Extract date columns from <th data-date-key="...">
        headers = table.find("thead")
        if not headers:
            return {}
        dates: list[str] = []
        for th in headers.find_all("th"):
            dk = th.get("data-date-key")
            if dk:
                dates.append(dk)  # Already YYYY-MM-DD format

        if not dates:
            return {}

        # Extract rows
        result: dict[str, list[tuple[str, float | None]]] = {}
        tbody = table.find("tbody")
        if not tbody:
            return {}

        for tr in tbody.find_all("tr", recursive=False):
            tds = tr.find_all("td")
            if not tds:
                continue
            # First td is the label (may contain a <button> for sub-items)
            label_td = tds[0]
            label = label_td.get_text(strip=True)
            if not label:
                continue

            values: list[tuple[str, float | None]] = []
            for i, date_key in enumerate(dates):
                td_idx = i + 1  # offset by label column
                if td_idx < len(tds):
                    raw = tds[td_idx].get_text(strip=True)
                    raw = raw.replace(",", "").replace("%", "").strip()
                    if not raw or raw == "—" or raw == "-":
                        values.append((date_key, None))
                    else:
                        try:
                            values.append((date_key, float(raw)))
                        except ValueError:
                            values.append((date_key, None))
                else:
                    values.append((date_key, None))
            result[label] = values

        return result

    def parse_quarterly_from_html(self, symbol: str, html: str) -> list[QuarterlyResult]:
        """Parse quarterly results from Screener.in HTML #quarters section."""
        soup = BeautifulSoup(html, "html.parser")
        data = self._parse_table_section(soup, "quarters")
        if not data:
            return []

        # Flexible label matching
        def _find(candidates: list[str]) -> list[tuple[str, float | None]]:
            for c in candidates:
                cl = c.lower()
                for label, vals in data.items():
                    if label.lower().rstrip("+") == cl:
                        return vals
            return []

        revenue_vals = _find(["Sales", "Revenue", "Total Revenue"])
        expenses_vals = _find(["Expenses"])
        op_income_vals = _find(["Operating Profit", "Financing Profit"])
        opm_vals = _find(["OPM %", "OPM%", "Financing Margin %", "Financing Margin%"])
        other_income_vals = _find(["Other Income"])
        interest_vals = _find(["Interest"])
        depreciation_vals = _find(["Depreciation"])
        pbt_vals = _find(["Profit before tax"])
        tax_pct_vals = _find(["Tax %", "Tax%"])
        net_income_vals = _find(["Net Profit"])
        eps_vals = _find(["EPS in Rs"])

        # Build date list from first non-empty series
        first_series = revenue_vals or expenses_vals or net_income_vals
        if not first_series:
            return []

        results = []
        for i, (date_key, _) in enumerate(first_series):
            def _val(series: list[tuple[str, float | None]]) -> float | None:
                return series[i][1] if i < len(series) else None

            revenue = _val(revenue_vals)
            operating_income = _val(op_income_vals)
            net_income = _val(net_income_vals)
            depreciation = _val(depreciation_vals)

            # OPM from Screener (already percentage, e.g. 30 means 30%)
            opm_raw = _val(opm_vals)
            operating_margin = opm_raw / 100 if opm_raw is not None else None

            # EBITDA = operating_income + depreciation
            ebitda = None
            if operating_income is not None and depreciation is not None:
                ebitda = operating_income + depreciation

            # Net margin
            net_margin = None
            if net_income is not None and revenue and revenue != 0:
                net_margin = net_income / revenue

            results.append(QuarterlyResult(
                symbol=symbol,
                quarter_end=date_key,
                revenue=revenue,
                expenses=_val(expenses_vals),
                operating_income=operating_income,
                net_income=net_income,
                ebitda=ebitda,
                eps=_val(eps_vals),
                operating_margin=operating_margin,
                net_margin=net_margin,
                other_income=_val(other_income_vals),
                depreciation=depreciation,
                interest=_val(interest_vals),
                profit_before_tax=_val(pbt_vals),
                tax_pct=_val(tax_pct_vals),
            ))

        results.sort(key=lambda r: r.quarter_end)
        return results

    def parse_ratios_from_html(self, symbol: str, html: str) -> list[ScreenerRatios]:
        """Parse efficiency ratios from Screener.in HTML #ratios section."""
        soup = BeautifulSoup(html, "html.parser")
        data = self._parse_table_section(soup, "ratios")
        if not data:
            return []

        def _find(candidates: list[str]) -> list[tuple[str, float | None]]:
            for c in candidates:
                cl = c.lower()
                for label, vals in data.items():
                    if label.lower().rstrip("+") == cl:
                        return vals
            return []

        debtor_vals = _find(["Debtor Days"])
        inventory_vals = _find(["Inventory Days"])
        payable_vals = _find(["Days Payable"])
        ccc_vals = _find(["Cash Conversion Cycle"])
        wc_vals = _find(["Working Capital Days"])
        roce_vals = _find(["ROCE %", "ROCE%", "ROE %", "ROE%"])

        first_series = debtor_vals or roce_vals or wc_vals
        if not first_series:
            return []

        results = []
        for i, (date_key, _) in enumerate(first_series):
            def _val(series: list[tuple[str, float | None]]) -> float | None:
                return series[i][1] if i < len(series) else None

            results.append(ScreenerRatios(
                symbol=symbol,
                fiscal_year_end=date_key,
                debtor_days=_val(debtor_vals),
                inventory_days=_val(inventory_vals),
                days_payable=_val(payable_vals),
                cash_conversion_cycle=_val(ccc_vals),
                working_capital_days=_val(wc_vals),
                roce_pct=_val(roce_vals),
            ))

        results.sort(key=lambda r: r.fiscal_year_end)
        return results

    def parse_documents_from_html(self, html: str) -> dict:
        """Parse #documents section for concall and annual report URLs.

        Returns:
            {
                "concalls": [
                    {"quarter": "Jan 2026", "transcript_url": "...", "ppt_url": "...", "recording_url": "..."},
                    ...
                ],
                "annual_reports": [
                    {"year": "Financial Year 2025", "url": "..."},
                    ...
                ]
            }
        """
        soup = BeautifulSoup(html, "html.parser")
        section = soup.find("section", id="documents")
        if not section:
            return {"concalls": [], "annual_reports": []}

        result: dict = {"concalls": [], "annual_reports": []}

        # Find all h3 headers in the documents section
        for h3 in section.find_all("h3"):
            heading_text = h3.get_text(strip=True).lower()

            if "concall" in heading_text:
                ul = h3.find_next("ul", class_="list-links")
                if not ul:
                    continue
                for li in ul.find_all("li"):
                    # Quarter label from the div
                    quarter_div = li.find("div")
                    quarter = quarter_div.get_text(strip=True) if quarter_div else ""

                    transcript_url = ""
                    ppt_url = ""
                    recording_url = ""

                    # Only scrape <a> tags with class concall-link (skip <button>)
                    for a_tag in li.find_all("a", class_="concall-link"):
                        link_text = a_tag.get_text(strip=True).lower()
                        href = a_tag.get("href", "")
                        if not href:
                            continue
                        if "transcript" in link_text:
                            transcript_url = href
                        elif "ppt" in link_text:
                            ppt_url = href
                        elif "rec" in link_text:
                            recording_url = href

                    if quarter:
                        result["concalls"].append({
                            "quarter": quarter,
                            "transcript_url": transcript_url,
                            "ppt_url": ppt_url,
                            "recording_url": recording_url,
                        })

            elif "annual report" in heading_text:
                ul = h3.find_next("ul", class_="list-links")
                if not ul:
                    continue
                for li in ul.find_all("li"):
                    a_tag = li.find("a")
                    if not a_tag:
                        continue
                    # The year label is the text content (without the nested div text)
                    # Clone and remove nested elements to get clean text
                    year_text = a_tag.get_text(strip=True)
                    # Remove the "from bse" suffix if present
                    nested_div = a_tag.find("div")
                    if nested_div:
                        nested_text = nested_div.get_text(strip=True)
                        year_text = year_text.replace(nested_text, "").strip()
                    href = a_tag.get("href", "")
                    if year_text and href:
                        result["annual_reports"].append({
                            "year": year_text,
                            "url": href,
                        })

        return result

    def parse_growth_rates_from_html(self, html: str) -> dict:
        """Parse compounded growth rates tables from Screener HTML.

        Returns dict like:
        {
            "sales_10y": 0.23, "sales_5y": 0.17, "sales_3y": 0.23, "sales_ttm": 0.13,
            "profit_10y": 0.30, ...,
            "price_10y": None, "price_5y": -0.13, ...,
            "roe_10y": None, "roe_5y": 0.20, ...
        }
        """
        soup = BeautifulSoup(html, "html.parser")
        tables = soup.find_all("table", class_="ranges-table")
        result: dict = {}

        key_map = {
            "Compounded Sales Growth": "sales",
            "Compounded Profit Growth": "profit",
            "Stock Price CAGR": "price",
            "Return on Equity": "roe",
        }
        period_map = {
            "10 Years:": "10y",
            "5 Years:": "5y",
            "3 Years:": "3y",
            "TTM:": "ttm",
            "1 Year:": "1y",
            "Last Year:": "last",
        }

        for table in tables:
            header = table.find("th")
            if not header:
                continue
            name = header.get_text(strip=True)
            prefix = key_map.get(name)
            if not prefix:
                continue

            for row in table.find_all("tr")[1:]:
                cells = row.find_all("td")
                if len(cells) != 2:
                    continue
                period_text = cells[0].get_text(strip=True)
                value_text = cells[1].get_text(strip=True).replace("%", "").strip()
                suffix = period_map.get(period_text)
                if not suffix:
                    continue
                try:
                    result[f"{prefix}_{suffix}"] = float(value_text) / 100
                except (ValueError, TypeError):
                    result[f"{prefix}_{suffix}"] = None

        # Screener uses "Last Year" for ROE which is actually the latest FY
        # Also add roe_3y alias if only "Last Year" was parsed
        if "roe_3y" in result and "Last Year:" in str(tables):
            # The Last Year mapped to 3y, but we also want it as a separate key
            pass

        return result

    @staticmethod
    def _parse_quarter_key(q: str) -> str:
        """Convert 'Dec 2025' to '2025-12' for proper chronological sorting."""
        month_map = {
            "Jan": "01", "Feb": "02", "Mar": "03", "Apr": "04",
            "May": "05", "Jun": "06", "Jul": "07", "Aug": "08",
            "Sep": "09", "Oct": "10", "Nov": "11", "Dec": "12",
        }
        parts = q.split()
        if len(parts) == 2 and parts[0] in month_map:
            return f"{parts[1]}-{month_map[parts[0]]}"
        return q

    def fetch_shareholder_details(self, html: str) -> dict:
        """Fetch detailed shareholder data from Screener's investor API.

        Fetches both quarterly and yearly views for DII and FII.
        Returns dict with keys: dii, fii (quarterly), dii_yearly, fii_yearly.
        Each is a list of {name, quarters: {quarter: pct}, url, all_quarters (sorted)}.
        """
        soup = BeautifulSoup(html, "html.parser")
        info_el = soup.find(attrs={"data-company-id": True})
        if not info_el:
            return {}

        company_id = info_el["data-company-id"]
        result: dict = {}

        api_calls = {
            "dii": ("domestic_institutions", "quarterly"),
            "fii": ("foreign_institutions", "quarterly"),
            "dii_yearly": ("domestic_institutions", "yearly"),
            "fii_yearly": ("foreign_institutions", "yearly"),
        }

        for key, (classification, period) in api_calls.items():
            try:
                resp = self._client.get(
                    f"{_SCREENER_BASE}/api/3/{company_id}/investors/{classification}/{period}/"
                )
                if resp.status_code != 200:
                    continue

                raw = resp.json()
                entities = []
                all_quarters: set[str] = set()

                for name, data in raw.items():
                    attrs = data.pop("setAttributes", {})
                    url = attrs.get("data-person-url", "")
                    quarters = {k: float(v) for k, v in data.items()}
                    all_quarters.update(quarters.keys())
                    if quarters:
                        # Find latest quarter by proper date sorting
                        sorted_qs = sorted(quarters.keys(), key=self._parse_quarter_key, reverse=True)
                        entities.append({
                            "name": name,
                            "latest_pct": quarters[sorted_qs[0]],
                            "latest_quarter": sorted_qs[0],
                            "quarters": quarters,
                            "url": url,
                        })

                entities.sort(key=lambda x: x["latest_pct"], reverse=True)

                # Sort all quarters chronologically (newest first)
                sorted_all = sorted(all_quarters, key=self._parse_quarter_key, reverse=True)
                result[key] = entities
                result[f"{key}_quarters"] = sorted_all
            except Exception:
                continue

        return result

    def _get_company_id(self, html: str) -> str | None:
        """Extract the Screener internal company ID from page HTML.

        Used for the /api/company/{id}/chart/ endpoint.
        """
        match = re.search(r"/api/company/(\d+)/", html)
        return match.group(1) if match else None

    def fetch_chart_data(self, symbol: str, html: str | None = None) -> dict:
        """Fetch ALL 6 chart datasets from Screener's chart API in one call.

        For fetching a single chart type, use fetch_chart_data_by_type() instead.

        Returns dict with keys: price_chart, pe_chart, margins_chart,
        ev_ebitda_chart, pb_chart, mcap_sales_chart.
        """
        if html is None:
            html = self.fetch_company_page(symbol)

        company_id = self._get_company_id(html)
        if not company_id:
            return {}

        result = {}
        queries = {
            "price_chart": "Price-DMA50-DMA200-Volume",
            "pe_chart": "Price+to+Earning-Median+PE-EPS",
            "margins_chart": "GPM-OPM-NPM-Quarter+Sales",
            "ev_ebitda_chart": "EV+Multiple-Median+EV+Multiple-EBITDA",
            "pb_chart": "Price+to+book+value-Median+PBV-Book+value",
            "mcap_sales_chart": "Market+Cap+to+Sales-Median+Market+Cap+to+Sales-Sales",
        }

        for key, q in queries.items():
            try:
                resp = self._client.get(
                    f"{_SCREENER_BASE}/api/company/{company_id}/chart/?q={q}&days=10000"
                )
                if resp.status_code == 200:
                    data = resp.json()
                    datasets = {}
                    for ds in data.get("datasets", []):
                        label = ds.get("label", "")
                        datasets[label] = ds.get("values", [])
                    result[key] = datasets
            except Exception:
                continue

        return result

    # --- Phase 2: New Screener API methods for research agent ---

    def _get_both_ids(self, html: str) -> tuple[str, str]:
        """Extract both company_id and warehouse_id from company page HTML."""
        soup = BeautifulSoup(html, "html.parser")
        info_el = soup.find(id="company-info") or soup.find(attrs={"data-company-id": True})
        company_id = ""
        warehouse_id = ""
        if info_el:
            company_id = info_el.get("data-company-id", "")
            warehouse_id = info_el.get("data-warehouse-id", "")
        if not company_id:
            m = re.search(r'data-company-id="(\d+)"', html)
            company_id = m.group(1) if m else ""
        if not warehouse_id:
            m = re.search(r'formaction="/user/company/export/(\d+)/"', html)
            warehouse_id = m.group(1) if m else ""
        return company_id, warehouse_id

    def search(self, query: str) -> list[dict]:
        """Search for companies. Returns [{id, name, url}, ...]."""
        resp = self._client.get(f"{_SCREENER_BASE}/api/company/search/", params={"q": query})
        resp.raise_for_status()
        return resp.json()

    def fetch_chart_data_by_type(self, company_id: str, chart_type: str, days: int = 10000) -> dict:
        """Fetch a single chart type from Chart API.

        chart_type: 'price', 'pe', 'sales_margin', 'ev_ebitda', 'pbv', 'mcap_sales'
        Returns: {"datasets": [{"metric": str, "label": str, "values": [[date, value], ...]}]}
        """
        queries = {
            "price": "Price-DMA50-DMA200-Volume",
            "pe": "Price+to+Earning-Median+PE-EPS",
            "sales_margin": "GPM-OPM-NPM-Quarter+Sales",
            "ev_ebitda": "EV+Multiple-Median+EV+Multiple-EBITDA",
            "pbv": "Price+to+book+value-Median+PBV-Book+value",
            "mcap_sales": "Market+Cap+to+Sales-Median+Market+Cap+to+Sales-Sales",
        }
        q = queries.get(chart_type)
        if not q:
            return {"datasets": []}
        url = f"{_SCREENER_BASE}/api/company/{company_id}/chart/?q={q}&days={days}&consolidated=true"
        resp = self._client.get(url)
        resp.raise_for_status()
        return resp.json()

    # Deprecated alias
    fetch_chart_data_single = fetch_chart_data_by_type

    def fetch_peers(self, warehouse_id: str) -> list[dict]:
        """Fetch peer comparison table. Returns list of peer dicts."""
        url = f"{_SCREENER_BASE}/api/company/{warehouse_id}/peers/"
        resp = self._client.get(url)
        resp.raise_for_status()
        soup = BeautifulSoup(resp.text, "html.parser")
        table = soup.find("table")
        if not table:
            return []
        headers = [th.get_text(strip=True) for th in table.find_all("th")]
        peers = []
        for tr in table.find_all("tr")[1:]:
            tds = tr.find_all("td")
            if len(tds) < 2:
                continue
            row: dict = {}
            for i, td in enumerate(tds):
                if i < len(headers):
                    key = headers[i].lower().replace(" ", "_").replace(".", "").replace("(", "").replace(")", "").replace("%", "pct")
                    val = td.get_text(strip=True).replace(",", "")
                    try:
                        row[key] = float(val)
                    except ValueError:
                        row[key] = val
                    a = td.find("a")
                    if a:
                        row["url"] = a.get("href", "")
            peers.append(row)
        return peers

    def fetch_shareholders(self, company_id: str) -> dict[str, list[dict]]:
        """Fetch individual shareholder details for each category from API.

        Returns: {"promoters": [{name, values: {quarter: pct}, url}, ...], ...}
        """
        import time as _time
        classifications = ["promoters", "foreign_institutions", "domestic_institutions", "public"]
        result: dict[str, list[dict]] = {}
        for cls in classifications:
            url = f"{_SCREENER_BASE}/api/3/{company_id}/investors/{cls}/quarterly/"
            try:
                resp = self._client.get(url)
                resp.raise_for_status()
                raw = resp.json()
                holders: list[dict] = []
                if isinstance(raw, dict):
                    for name, data in raw.items():
                        if isinstance(data, dict):
                            person_url = data.pop("data-person-url", data.pop("setAttributes", {}).get("data-person-url", ""))
                            holders.append({"name": name, "values": data, "url": person_url})
                result[cls] = holders
            except Exception:
                result[cls] = []
            _time.sleep(1)
        return result

    def fetch_schedules(self, company_id: str, section: str, parent: str) -> dict:
        """Fetch schedule (sub-item breakdown) for a specific line item.

        section: 'quarters', 'profit-loss', 'balance-sheet', 'cash-flow'
        parent: e.g., 'Sales', 'Expenses', 'Borrowings'
        """
        url = f"{_SCREENER_BASE}/api/company/{company_id}/schedules/"
        params = {"parent": parent, "section": section, "consolidated": ""}
        resp = self._client.get(url, params=params)
        resp.raise_for_status()
        data = resp.json()
        return data if isinstance(data, dict) else {}

    def fetch_all_schedules(self, company_id: str) -> dict[str, dict[str, dict]]:
        """Fetch all schedule breakdowns across all sections."""
        import time as _time
        schedule_parents = {
            "quarters": ["Sales", "Expenses", "Other Income", "Net Profit"],
            "profit-loss": ["Sales", "Expenses", "Other Income", "Net Profit"],
            "balance-sheet": ["Borrowings", "Other Liabilities", "Fixed Assets", "Other Assets"],
            "cash-flow": ["Cash from Operating Activity", "Cash from Investing Activity", "Cash from Financing Activity"],
        }
        result: dict[str, dict[str, dict]] = {}
        for section, parents in schedule_parents.items():
            result[section] = {}
            for parent in parents:
                try:
                    result[section][parent] = self.fetch_schedules(company_id, section, parent)
                except Exception:
                    result[section][parent] = {}
                _time.sleep(1)
        return result

    def close(self) -> None:
        self._client.close()

    def __enter__(self) -> ScreenerClient:
        return self

    def __exit__(self, *args: object) -> None:
        self.close()
