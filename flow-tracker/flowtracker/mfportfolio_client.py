"""MF portfolio disclosure client — downloads and parses AMC monthly XLSX files."""

from __future__ import annotations

import io
import logging
import re
import tempfile
import zipfile
from calendar import monthrange
from datetime import date

import httpx

from flowtracker.mfportfolio_models import MFSchemeHolding

logger = logging.getLogger(__name__)

_UA = "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36"


class MFPortfolioClient:
    """Client for downloading and parsing MF scheme portfolio disclosures."""

    def __init__(self) -> None:
        self._client = httpx.Client(
            timeout=httpx.Timeout(connect=15.0, read=120.0, write=10.0, pool=10.0),
            follow_redirects=True,
            headers={"User-Agent": _UA},
        )

    def fetch_amc(self, amc: str, month: str) -> list[MFSchemeHolding]:
        """Fetch holdings for a specific AMC and month.

        Args:
            amc: AMC code — "SBI", "ICICI", "PPFAS", "QUANT", "UTI"
            month: Month string — "2026-02"

        Returns list of MFSchemeHolding.
        """
        fetchers = {
            "SBI": self._fetch_sbi,
            "ICICI": self._fetch_icici,
            "PPFAS": self._fetch_ppfas,
            "QUANT": self._fetch_quant,
            "UTI": self._fetch_uti,
        }
        fetcher = fetchers.get(amc.upper())
        if not fetcher:
            logger.warning("Unknown AMC: %s", amc)
            return []
        try:
            return fetcher(month)
        except Exception as e:
            logger.error("Failed to fetch %s for %s: %s", amc, month, e)
            return []

    def fetch_all(self, month: str) -> list[MFSchemeHolding]:
        """Fetch holdings from all supported AMCs for a month."""
        all_holdings: list[MFSchemeHolding] = []
        for amc in ["SBI", "ICICI", "PPFAS", "QUANT", "UTI"]:
            logger.info("Fetching %s for %s...", amc, month)
            holdings = self.fetch_amc(amc, month)
            all_holdings.extend(holdings)
            logger.info("  %s: %d holdings", amc, len(holdings))
        return all_holdings

    # -- SBI MF --

    def _fetch_sbi(self, month: str) -> list[MFSchemeHolding]:
        """Fetch SBI MF consolidated portfolio XLSX."""
        year, mon = _parse_month(month)
        last_day = monthrange(year, mon)[1]
        ordinal = _ordinal(last_day)
        month_name = date(year, mon, 1).strftime("%B").lower()

        url = (
            f"https://www.sbimf.com/docs/default-source/scheme-portfolios/"
            f"all-schemes-monthly-portfolio---as-on-{last_day}{ordinal}-{month_name}-{year}.xlsx"
        )
        return self._download_and_parse_xlsx(url, "SBI", month, multi_sheet=True)

    # -- ICICI Prudential --

    def _fetch_icici(self, month: str) -> list[MFSchemeHolding]:
        """Fetch ICICI Prudential portfolio ZIP."""
        year, mon = _parse_month(month)
        month_full = date(year, mon, 1).strftime("%B")

        # Month dir is inconsistent — try abbreviated first, then full
        month_abbr = date(year, mon, 1).strftime("%b")
        for month_dir in [month_abbr, month_full]:
            url = (
                f"https://www.icicipruamc.com/blob/downloads/Files/"
                f"Monthly%20Portfolio%20Disclosures/{year}/{month_dir}/"
                f"Monthly-Portfolio-Disclosure-{month_full}-{year}.zip"
            )
            try:
                resp = self._client.get(url)
                if resp.status_code == 200:
                    return self._parse_icici_zip(resp.content, month)
            except Exception:
                continue

        logger.warning("ICICI portfolio not found for %s", month)
        return []

    def _parse_icici_zip(self, content: bytes, month: str) -> list[MFSchemeHolding]:
        """Parse ICICI ZIP containing per-scheme XLSX files."""
        holdings: list[MFSchemeHolding] = []
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for name in zf.namelist():
                if not name.lower().endswith((".xlsx", ".xls")):
                    continue
                try:
                    data = zf.read(name)
                    scheme_holdings = self._parse_single_xlsx(
                        data, "ICICI", month, name,
                    )
                    holdings.extend(scheme_holdings)
                except Exception as e:
                    logger.debug("Skipping ICICI file %s: %s", name, e)
        return holdings

    # -- PPFAS --

    def _fetch_ppfas(self, month: str) -> list[MFSchemeHolding]:
        """Fetch PPFAS consolidated portfolio XLS."""
        year, mon = _parse_month(month)
        last_day = monthrange(year, mon)[1]
        month_name = date(year, mon, 1).strftime("%B")

        url = (
            f"https://amc.ppfas.com/downloads/portfolio-disclosure/{year}/"
            f"PPFAS_Monthly_Portfolio_Report_{month_name}_{last_day}_{year}.xls"
        )
        return self._download_and_parse_xlsx(url, "PPFAS", month, multi_sheet=True)

    # -- Quant MF --

    def _fetch_quant(self, month: str) -> list[MFSchemeHolding]:
        """Fetch Quant MF portfolio via AJAX discovery."""
        year, mon = _parse_month(month)

        # Call AJAX to discover filename
        try:
            resp = self._client.post(
                "https://quantmutual.com/statutorydisclosures.aspx/displaydisclouser",
                json={"id": str(year), "cat": "Monthly Portfolio"},
                headers={"Content-Type": "application/json"},
            )
            resp.raise_for_status()
            data = resp.json()
            html = data.get("d", "")

            # Find XLSX links in the HTML response
            month_abbr = date(year, mon, 1).strftime("%b").lower()
            month_full = date(year, mon, 1).strftime("%B").lower()
            year_short = str(year)[2:]

            # Look for links matching this month
            links = re.findall(r'href="([^"]+\.xlsx)"', html, re.IGNORECASE)
            if not links:
                links = re.findall(r'(Admin/disclouser/[^"\']+\.xlsx)', html, re.IGNORECASE)

            target_link = None
            for link in links:
                link_lower = link.lower()
                if (month_abbr in link_lower or month_full in link_lower or
                        f"{mon:02d}{year}" in link_lower or
                        f"{month_abbr}{year_short}" in link_lower):
                    target_link = link
                    break

            if not target_link:
                # Try first link as fallback for single-month responses
                if links:
                    target_link = links[0]
                else:
                    logger.warning("No Quant portfolio link found for %s", month)
                    return []

            # Build full URL
            if target_link.startswith("http"):
                url = target_link
            elif target_link.startswith("/"):
                url = f"https://quantmutual.com{target_link}"
            else:
                url = f"https://quantmutual.com/{target_link}"

            return self._download_and_parse_xlsx(url, "QUANT", month, multi_sheet=True)

        except Exception as e:
            logger.error("Quant AJAX fetch failed for %s: %s", month, e)
            return []

    # -- UTI MF --

    def _fetch_uti(self, month: str) -> list[MFSchemeHolding]:
        """Fetch UTI MF portfolio ZIP from CloudFront."""
        year, mon = _parse_month(month)
        last_day = monthrange(year, mon)[1]

        # Upload month is the following month
        next_mon = mon + 1
        next_year = year
        if next_mon > 12:
            next_mon = 1
            next_year += 1

        month_name = date(next_year, next_mon, 1).strftime("%B")

        # Try multiple URL patterns (inconsistent naming)
        patterns = [
            f"https://d3ce1o48hc5oli.cloudfront.net/static/generic-zip/{month_name}-{str(next_year)[2:]}/FW_%20UTI_MF_Scheme_portfolios-{last_day:02d}.{mon:02d}.{year}.zip",
            f"https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/{next_year}-{next_mon:02d}/fw_uti_mf_scheme_portfolios_{last_day:02d}.{mon:02d}.{year}_1.zip",
            f"https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/{next_year}-{next_mon:02d}/fw_uti_mf_scheme_portfolios_{last_day:02d}.{mon:02d}.{year}_0.zip",
            f"https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/{next_year}-{next_mon:02d}/fw_uti_mf_portfolios_{last_day:02d}.{mon:02d}.{year}_1.zip",
            f"https://d3ce1o48hc5oli.cloudfront.net/s3fs-public/{next_year}-{next_mon:02d}/fw_uti_mf_portfolios_{last_day:02d}.{mon:02d}.{year}_0.zip",
        ]

        for url in patterns:
            try:
                resp = self._client.get(url)
                if resp.status_code == 200 and len(resp.content) > 1000:
                    return self._parse_uti_zip(resp.content, month)
            except Exception:
                continue

        logger.warning("UTI portfolio not found for %s", month)
        return []

    def _parse_uti_zip(self, content: bytes, month: str) -> list[MFSchemeHolding]:
        """Parse UTI ZIP, find the SEBI Exposure XLSX."""
        holdings: list[MFSchemeHolding] = []
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for name in zf.namelist():
                if "sebi" in name.lower() and "exposure" in name.lower():
                    data = zf.read(name)
                    holdings = self._parse_single_xlsx(data, "UTI", month, name)
                    break
        return holdings

    # -- Common XLSX Parsing --

    def _download_and_parse_xlsx(
        self, url: str, amc: str, month: str, multi_sheet: bool = False,
    ) -> list[MFSchemeHolding]:
        """Download an XLSX/XLS file and parse equity holdings."""
        resp = self._client.get(url)
        resp.raise_for_status()
        content = resp.content

        # Detect format: .xls (legacy) vs .xlsx (OOXML)
        is_xls = url.lower().endswith(".xls") or (
            content[:4] != b"PK\x03\x04"  # XLSX starts with PK zip header
        )

        if is_xls:
            return self._parse_xls_file(content, amc, month)

        if multi_sheet:
            return self._parse_multi_sheet_xlsx(content, amc, month)
        return self._parse_single_xlsx(content, amc, month)

    def _parse_xls_file(
        self, content: bytes, amc: str, month: str,
    ) -> list[MFSchemeHolding]:
        """Parse a legacy .xls file using xlrd."""
        import xlrd
        holdings: list[MFSchemeHolding] = []
        wb = xlrd.open_workbook(file_contents=content)
        for sheet_idx in range(wb.nsheets):
            ws = wb.sheet_by_index(sheet_idx)
            sheet_holdings = self._extract_equity_holdings_xlrd(
                ws, amc, month, ws.name,
            )
            holdings.extend(sheet_holdings)
        return holdings

    def _parse_multi_sheet_xlsx(
        self, content: bytes, amc: str, month: str,
    ) -> list[MFSchemeHolding]:
        """Parse an XLSX with multiple scheme sheets."""
        import openpyxl

        holdings: list[MFSchemeHolding] = []
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_holdings = self._extract_equity_holdings(
                ws, amc, month, sheet_name,
            )
            holdings.extend(sheet_holdings)

        wb.close()
        return holdings

    def _parse_single_xlsx(
        self, content: bytes, amc: str, month: str, filename: str = "",
    ) -> list[MFSchemeHolding]:
        """Parse a single-sheet XLSX file."""
        import openpyxl

        holdings: list[MFSchemeHolding] = []
        try:
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                scheme = filename.replace(".xlsx", "").replace(".xls", "") or sheet_name
                # Clean up scheme name
                scheme = re.sub(r'^ICICI Prudential ', '', scheme)
                sheet_holdings = self._extract_equity_holdings(ws, amc, month, scheme)
                holdings.extend(sheet_holdings)
            wb.close()
        except Exception as e:
            # Try xlrd for .xls files
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=content)
                for sheet_idx in range(wb.nsheets):
                    ws = wb.sheet_by_index(sheet_idx)
                    scheme = ws.name
                    sheet_holdings = self._extract_equity_holdings_xlrd(
                        ws, amc, month, scheme,
                    )
                    holdings.extend(sheet_holdings)
            except ImportError:
                logger.warning("xlrd not installed, can't parse .xls files: %s", e)
            except Exception as e2:
                logger.warning("Failed to parse file: %s / %s", e, e2)

        return holdings

    def _extract_equity_holdings(
        self, ws: object, amc: str, month: str, scheme_name: str,
    ) -> list[MFSchemeHolding]:
        """Extract equity holdings from an openpyxl worksheet."""
        holdings: list[MFSchemeHolding] = []
        in_equity_section = False
        header_row: dict[str, int] = {}

        for row in ws.iter_rows(values_only=False):
            cells = [c.value for c in row]

            # Skip empty rows
            if not any(cells):
                continue

            first = str(cells[0] or "").strip().upper()

            # Detect equity section
            if "EQUITY" in first and "RELATED" in first:
                in_equity_section = True
                continue

            # Detect end of equity section
            if in_equity_section and any(
                kw in first for kw in ["DEBT", "MONEY MARKET", "TOTAL", "NET ASSETS", "GRAND TOTAL"]
            ):
                if "TOTAL" in first and "EQUITY" not in first:
                    in_equity_section = False
                    continue

            # Detect header row
            cell_strs = [str(c or "").strip().upper() for c in cells]
            if "ISIN" in cell_strs:
                header_row = {}
                for i, val in enumerate(cell_strs):
                    if "ISIN" in val:
                        header_row["isin"] = i
                    elif "NAME" in val and "INSTRUMENT" in val:
                        header_row["name"] = i
                    elif val == "QUANTITY" or "QUANTITY" in val:
                        header_row["qty"] = i
                    elif "MARKET" in val and ("VALUE" in val or "FAIR" in val):
                        header_row["value"] = i
                    elif "% TO" in val and ("NAV" in val or "AUM" in val or "NET" in val):
                        header_row["pct"] = i
                    elif val == "INDUSTRY" or "INDUSTRY" in val:
                        header_row["industry"] = i
                continue

            if not header_row or "isin" not in header_row:
                continue

            # Parse data row
            holding = self._parse_holding_row(cells, header_row, amc, month, scheme_name)
            if holding:
                holdings.append(holding)

        return holdings

    def _extract_equity_holdings_xlrd(
        self, ws: object, amc: str, month: str, scheme_name: str,
    ) -> list[MFSchemeHolding]:
        """Extract equity holdings from an xlrd worksheet."""
        holdings: list[MFSchemeHolding] = []
        header_row: dict[str, int] = {}

        for row_idx in range(ws.nrows):
            cells = [ws.cell_value(row_idx, col) for col in range(ws.ncols)]

            if not any(cells):
                continue

            cell_strs = [str(c or "").strip().upper() for c in cells]

            # Detect header
            if "ISIN" in cell_strs:
                header_row = {}
                for i, val in enumerate(cell_strs):
                    if "ISIN" in val:
                        header_row["isin"] = i
                    elif "NAME" in val:
                        header_row["name"] = i
                    elif "QUANTITY" in val:
                        header_row["qty"] = i
                    elif "MARKET" in val or "FAIR" in val:
                        header_row["value"] = i
                    elif "% TO" in val:
                        header_row["pct"] = i
                continue

            if not header_row or "isin" not in header_row:
                continue

            holding = self._parse_holding_row(cells, header_row, amc, month, scheme_name)
            if holding:
                holdings.append(holding)

        return holdings

    def _parse_holding_row(
        self, cells: list, header: dict[str, int], amc: str, month: str, scheme: str,
    ) -> MFSchemeHolding | None:
        """Parse a single data row into a MFSchemeHolding."""
        try:
            isin = str(cells[header["isin"]] or "").strip()
            if not isin or len(isin) < 10 or not isin.startswith("IN"):
                return None

            name_idx = header.get("name", header.get("isin", 0) - 1)
            stock_name = str(cells[name_idx] or "").strip()
            if not stock_name:
                return None

            qty_idx = header.get("qty")
            qty = 0
            if qty_idx is not None and cells[qty_idx]:
                try:
                    qty = int(float(str(cells[qty_idx]).replace(",", "")))
                except (ValueError, TypeError):
                    qty = 0

            value_idx = header.get("value")
            value = 0.0
            if value_idx is not None and cells[value_idx]:
                try:
                    value = float(str(cells[value_idx]).replace(",", ""))
                except (ValueError, TypeError):
                    value = 0.0

            pct_idx = header.get("pct")
            pct = 0.0
            if pct_idx is not None and cells[pct_idx]:
                try:
                    pct = float(str(cells[pct_idx]).replace(",", "").replace("%", ""))
                except (ValueError, TypeError):
                    pct = 0.0

            if qty == 0 and value == 0:
                return None

            return MFSchemeHolding(
                month=month,
                amc=amc,
                scheme_name=scheme[:100],
                isin=isin,
                stock_name=stock_name[:100],
                quantity=qty,
                market_value_cr=value / 100,
                pct_of_nav=pct,
            )
        except (IndexError, KeyError):
            return None

    def close(self) -> None:
        self._client.close()

    def __enter__(self) -> MFPortfolioClient:
        return self

    def __exit__(self, *args: object) -> None:
        self.close()


def _parse_month(month: str) -> tuple[int, int]:
    """Parse '2026-02' into (2026, 2)."""
    parts = month.split("-")
    return int(parts[0]), int(parts[1])


def _ordinal(n: int) -> str:
    """Return English ordinal suffix for a number."""
    if 11 <= (n % 100) <= 13:
        return "th"
    return {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
